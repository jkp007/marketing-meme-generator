# ============================================================================
# SECTION 1: IMPORTS AND SETUP
# ============================================================================
import os
import pandas as pd
import sys
import time
import mimetypes
from datetime import datetime
from PIL import Image
from io import BytesIO
from IPython.display import display, Markdown, HTML
import streamlit as st
print("✅ Standard libraries imported")

# Import Google Gemini
from google import genai
from google.genai import types
print("✅ Google Gemini SDK imported")

# ============================================================================
# SECTION 2: API KEY CONFIGURATION (Using Kaggle Secrets)
# ============================================================================

print("\n" + "=" * 80)
print("🔑 CONFIGURING API KEY")
print("=" * 80)

# Get API key from environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
    GOOGLE_API_KEY = os.getenv("gemini_api_key")
    if GOOGLE_API_KEY:
        print("✅ API key loaded from environment variables")
        print(f"   Key length: {len(GOOGLE_API_KEY)} characters")
    else:
        raise ValueError("No API key found")
except Exception as e:
    print(f"❌ Error loading API key: {e}")
    print("\n📝 INSTRUCTIONS:")
    print("1. Get your FREE API key at: https://ai.studio/banana")
    print("2. In Kaggle, go to 'Add-ons' -> 'Secrets'")
    print("3. Add a new secret named 'gemini_api_key' with your API key")
    raise

# ============================================================================
# SECTION 3: INITIALIZE CLIENT
# ============================================================================

print("\n" + "=" * 80)
print("🚀 INITIALIZING CLIENT")
print("=" * 80)

# Initialize the client
client = genai.Client(api_key=GOOGLE_API_KEY)
MODEL_ID = "gemini-2.5-flash-image-preview"
MODEL_TEXT = "gemini-2.5-flash"

print(f"✅ Client initialized")
print(f"📷 Image Model: {MODEL_ID}")
print(f"📝 Text Model: {MODEL_TEXT}")

# Test the client
try:
    test_response = client.models.generate_content(
        model=MODEL_TEXT,
        contents="Say 'OK' if you're working"
    )
    print(f"✅ Test response: {test_response.text}")
    print("✅ Client test successful!")
except Exception as e:
    print(f"❌ Client test failed: {e}")

# ============================================================================
# SECTION 4: HELPER FUNCTIONS (Based on Working Example)
# ============================================================================

print("\n" + "=" * 80)
print("🛠️ SETTING UP HELPER FUNCTIONS")
print("=" * 80)

def save_binary_file(file_name, data):
    """Save binary data to file"""
    with open(file_name, "wb") as f:
        f.write(data)
    print(f"💾 File saved to: {file_name}")
    return file_name

def display_response(response):
    """Display text and images from response"""
    if not response:
        print("⚠️ No response to display")
        return
    
    # Handle candidates structure
    if hasattr(response, 'candidates') and response.candidates:
        for candidate in response.candidates:
            if hasattr(candidate, 'content') and candidate.content:
                if hasattr(candidate.content, 'parts'):
                    for part in candidate.content.parts:
                        process_part(part)
    # Handle direct parts structure
    elif hasattr(response, 'parts'):
        for part in response.parts:
            process_part(part)

def process_part(part):
    """Process a single part (text or image)"""
    # Display text
    if hasattr(part, 'text') and part.text:
        display(Markdown(part.text))
    
    # Display image
    if hasattr(part, 'inline_data') and part.inline_data:
        if hasattr(part.inline_data, 'data') and part.inline_data.data:
            try:
                image = Image.open(BytesIO(part.inline_data.data))
                print(f"📷 Image: {image.size[0]}x{image.size[1]} pixels")
                display(image)
            except Exception as e:
                print(f"⚠️ Could not display image: {e}")

def extract_and_save_images(response, prefix="image"):
    """Extract and save all images from response"""
    saved_files = []
    
    if not response:
        return saved_files
    
    # Handle candidates structure
    if hasattr(response, 'candidates') and response.candidates:
        for candidate in response.candidates:
            if hasattr(candidate, 'content') and candidate.content:
                if hasattr(candidate.content, 'parts'):
                    for i, part in enumerate(candidate.content.parts):
                        if hasattr(part, 'inline_data') and part.inline_data:
                            if hasattr(part.inline_data, 'data') and part.inline_data.data:
                                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                mime_type = getattr(part.inline_data, 'mime_type', 'image/png')
                                ext = mimetypes.guess_extension(mime_type) or '.png'
                                filename = f"{prefix}_{timestamp}_{i}{ext}"
                                save_binary_file(filename, part.inline_data.data)
                                saved_files.append(filename)
    
    # Handle direct parts structure
    elif hasattr(response, 'parts'):
        for i, part in enumerate(response.parts):
            if hasattr(part, 'inline_data') and part.inline_data:
                if hasattr(part.inline_data, 'data') and part.inline_data.data:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    mime_type = getattr(part.inline_data, 'mime_type', 'image/png')
                    ext = mimetypes.guess_extension(mime_type) or '.png'
                    filename = f"{prefix}_{timestamp}_{i}{ext}"
                    save_binary_file(filename, part.inline_data.data)
                    saved_files.append(filename)
    
    return saved_files

print("✅ Helper functions ready")

# ============================================================================
# SECTION 5: BASIC IMAGE GENERATION (NON-STREAMING)
# ============================================================================

print("\n" + "=" * 80)
print("🎨 DEMO 1: BASIC IMAGE GENERATION")
print("=" * 80)

def generate_image_basic(prompt):
    """Generate image using basic API call"""
    print(f"\n📝 Prompt: {prompt[:100]}...")
    
    try:
        response = client.models.generate_content(
            model=MODEL_ID,
            contents=[prompt],
            config=types.GenerateContentConfig(
                response_modalities=['TEXT', 'IMAGE']
            )
        )
        
        if response:
            print("✅ Response received!")
            display_response(response)
            saved_files = extract_and_save_images(response, "basic")
            return saved_files
        else:
            print("⚠️ Empty response")
            return []
            
    except Exception as e:
        print(f"❌ Error: {e}")
        return []
    
# ============================================================================
# SECTION 6: STREAMING IMAGE GENERATION (RECOMMENDED)
# ============================================================================

print("\n" + "=" * 80)
print("🎨 DEMO 2: STREAMING IMAGE GENERATION")
print("=" * 80)

def generate_image_streaming(prompt, save_prefix="stream"):
    """Generate image using streaming API (more reliable)"""
    print(f"\n📝 Prompt: {prompt[:100]}...")
    saved_files = []
    
    try:
        contents = [
            types.Content(
                role="user",
                parts=[types.Part.from_text(text=prompt)],
            )
        ]
        
        config = types.GenerateContentConfig(
            response_modalities=["IMAGE", "TEXT"]
        )
        
        print("🔄 Starting stream...")
        file_index = 0
        
        for chunk in client.models.generate_content_stream(
            model=MODEL_ID,
            contents=contents,
            config=config,
        ):
            print(".", end="", flush=True)
            
            if (chunk.candidates is None or 
                chunk.candidates[0].content is None or 
                chunk.candidates[0].content.parts is None):
                continue
            
            for part in chunk.candidates[0].content.parts:
                # Handle text
                if hasattr(part, 'text') and part.text:
                    print(f"\n📝 Text: {part.text[:100]}...")
                
                # Handle image
                if part.inline_data and part.inline_data.data:
                    mime_type = getattr(part.inline_data, 'mime_type', 'image/png')
                    ext = mimetypes.guess_extension(mime_type) or '.png'
                    filename = f"{save_prefix}_{file_index}{ext}"
                    save_binary_file(filename, part.inline_data.data)
                    saved_files.append(filename)
                    file_index += 1
                    
                    # Display the image
                    try:
                        image = Image.open(BytesIO(part.inline_data.data))
                        print(f"\n📷 Image generated: {image.size[0]}x{image.size[1]}")
                        display(image)
                    except:
                        pass
        
        print(f"\n✅ Streaming complete! Generated {len(saved_files)} image(s)")
        return saved_files
        
    except Exception as e:
        print(f"\n❌ Streaming error: {e}")
        return saved_files

# ============================================================================
# SECTION 7: STREAMLIT UI INTEGRATION WITH STEP-BY-STEP FEATURES
# ============================================================================

# Streamlit page config
st.set_page_config(page_title="Complytics Meme Generator", page_icon="🖼️", layout="wide")

# Import additional libraries for Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# Tab structure for steps
tab1, tab2, tab3 = st.tabs(["1. Business Info", "2. Generate Marketing Data", "3. Generate Memes & Export"])

# Global variables for data
marketing_data = None
generated_images = []

with tab1:
    st.header("📝 Step 1: Enter Business Information")
    business_name = st.text_input("Business Name", placeholder="e.g., Complytics.ai")
    website = st.text_input("Website", placeholder="e.g., https://complytics.ai")
    about_business = st.text_area("About Business", placeholder="Describe your business in a few sentences...", height=150)
    
    if st.button("Proceed to Generate Marketing Data"):
        if not business_name or not website or not about_business:
            st.error("Please fill in all fields.")
        else:
            st.session_state.business_info = {
                "name": business_name,
                "website": website,
                "about": about_business
            }
            st.success("Business info saved! Go to Step 2.")

with tab2:
    st.header("🎨 Step 2: Generate Marketing Data CSV")
    if "business_info" not in st.session_state:
        st.warning("Please complete Step 1 first.")
    else:
        info = st.session_state.business_info
        generate_csv_btn = st.button("Generate Marketing Data CSV")
        
        if generate_csv_btn:
            with st.spinner("Generating marketing data using AI..."):
                # Use Gemini to generate CSV data
                csv_prompt = f"""
                Based on the following business details:
                - Business Name: {info['name']}
                - Website: {info['website']}
                - About: {info['about']}
                
                Generate 10 rows of marketing data in CSV format with columns: meme_template, prompt, company_background, marketing_message, call_to_action, target_audience, platform, theme.
                Each row should be a unique meme idea tailored to the business.
                Output only the CSV data without headers or extra text.
                """
                
                try:
                    response = client.models.generate_content(
                        model=MODEL_TEXT,
                        contents=csv_prompt
                    )
                    csv_content = response.text.strip()
                    
                    # Parse CSV content into DataFrame
                    from io import StringIO
                    marketing_data = pd.read_csv(StringIO(csv_content), header=None, names=['meme_template', 'prompt', 'company_background', 'marketing_message', 'call_to_action', 'target_audience', 'platform', 'theme'])
                    
                    st.session_state.marketing_data = marketing_data
                    st.success("✅ Marketing data generated!")
                    st.dataframe(marketing_data)
                    
                    # Save CSV
                    marketing_data.to_csv("generated_marketing_data.csv", index=False)
                    st.download_button("Download CSV", data=marketing_data.to_csv(index=False), file_name="generated_marketing_data.csv", mime="text/csv")
                except Exception as e:
                    st.error(f"❌ Failed to generate CSV: {e}")

with tab3:
    st.header("🖼️ Step 3: Generate Memes & Export to Excel")
    if "marketing_data" not in st.session_state:
        st.warning("Please complete Step 2 first.")
    else:
        marketing_data = st.session_state.marketing_data
        
        # Select rows to generate memes
        options_list = ["Select All"] + [f"Row {i}: {marketing_data.iloc[i]['meme_template']}" for i in range(len(marketing_data))]
        selected_options = st.multiselect("Select rows to generate memes", options=options_list)
        
        if "Select All" in selected_options:
            selected_indices = list(range(len(marketing_data)))
        else:
            selected_indices = [int(opt.split(":")[0].replace("Row ", "")) for opt in selected_options if opt != "Select All"]
        
        generate_memes_btn = st.button("Generate Memes for Selected Rows")
        
        if generate_memes_btn and selected_indices:
            generated_images = []
            for idx in selected_indices:
                row = marketing_data.iloc[idx]
                meme_prompt = f"""
                "Create a meme using the '{row['meme_template']}' style. 
                The meme should tell a short, funny, and relatable story targeting {row['target_audience']} on {row['platform']}.  

                Company Background : Complytics.ai, {row['company_background']}  
                Marketing Message : {row['marketing_message']}  
                Call to Action to use Complytics.ai : {row['call_to_action']}  

                Theme: {row['theme']}  

                Ensure the meme humor connects with the target audience while subtly highlighting the value of the company. 
                The final meme should be engaging, visually clear, and suitable for viral social media marketing.
                Avoid adding any company logo. Dont make spelling mistake"
                """
                
                with st.spinner(f"Generating meme for Row {idx}..."):
                    try:
                        saved_files = generate_image_streaming(meme_prompt, f"meme_{idx}")
                        if saved_files:
                            generated_images.append({
                                "row": idx,
                                "image_path": saved_files[0],
                                "meme_template": row['meme_template'],
                                "prompt": row['prompt'],
                                "company_background": row['company_background'],
                                "marketing_message": row['marketing_message'],
                                "call_to_action": row['call_to_action'],
                                "target_audience": row['target_audience'],
                                "platform": row['platform'],
                                "theme": row['theme']
                            })
                            st.image(Image.open(saved_files[0]), caption=f"Meme for Row {idx}")
                    except Exception as e:
                        st.error(f"Failed for Row {idx}: {e}")
            
            st.session_state.generated_images = generated_images
            st.success(f"✅ Generated {len(generated_images)} memes!")
        
        # Export to Excel
        if "generated_images" in st.session_state and st.button("Export to Excel"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Memes"
            
            # Headers
            headers = ['Row', 'Meme Template', 'Prompt', 'Company Background', 'Marketing Message', 'Call to Action', 'Target Audience', 'Platform', 'Theme', 'Image']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row_idx, item in enumerate(st.session_state.generated_images, 2):
                ws.cell(row=row_idx, column=1, value=item['row'])
                ws.cell(row=row_idx, column=2, value=item['meme_template'])
                ws.cell(row=row_idx, column=3, value=item['prompt'])
                ws.cell(row=row_idx, column=4, value=item['company_background'])
                ws.cell(row=row_idx, column=5, value=item['marketing_message'])
                ws.cell(row=row_idx, column=6, value=item['call_to_action'])
                ws.cell(row=row_idx, column=7, value=item['target_audience'])
                ws.cell(row=row_idx, column=8, value=item['platform'])
                ws.cell(row=row_idx, column=9, value=item['theme'])
                
                # Insert image
                img = XLImage(item['image_path'])
                img.width = 200
                img.height = 200
                ws.add_image(img, f"J{row_idx}")
            
            # Save and download
            wb.save("memes_export.xlsx")
            with open("memes_export.xlsx", "rb") as f:
                st.download_button("Download Excel", data=f, file_name="memes_export.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Footer
st.markdown("---")
st.caption("Powered by Google Gemini. Ensure your API key is set in environment variables.")
